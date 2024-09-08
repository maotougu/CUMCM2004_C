import pandas as pd
import pulp
import openpyxl

# 读取数据
def read_data():
    sheet1 = pd.read_excel('datas/附件2.xlsx', sheet_name='2023年统计的相关数据')  # 土地类型上的亩产量数据
    sheet1 = sheet1.dropna(subset=['地块类型']) # 删除底部说明
    sheet1['地块类型'] = sheet1['地块类型'].str.strip() # 去空格
    sheet1['作物名称'] = sheet1['作物名称'].str.strip() # 去空格

    # 补充智慧大棚第一季在sheet1中的数据
    ordinary_greenhouse = sheet1[(sheet1['地块类型'] == '普通大棚') & (sheet1['种植季次'] == '第一季')]
    smart_greenhouse = ordinary_greenhouse.copy()   # 创建省略的“智慧大棚”数据
    smart_greenhouse['地块类型'] = '智慧大棚'
    sheet1 = pd.concat([sheet1, smart_greenhouse], ignore_index=True)

    sheet2 = pd.read_excel('datas/附件2.xlsx', sheet_name='2023年的农作物种植情况')
    sheet2 = sheet2.fillna(method='ffill')
    sheet2['作物名称'] = sheet2['作物名称'].str.strip() # 去空格

    sheet3 = pd.read_excel('datas/附件1.xlsx', sheet_name='乡村的现有耕地')
    sheet3['地块类型'] = sheet3['地块类型'].str.strip() # 去空格

    # 将 sheet2 和 sheet3 合并，得到每个地块的土地类型
    land_types = pd.merge(sheet2, sheet3, left_on='种植地块', right_on='地块名称')
    land_types = land_types.drop(columns=['说明 ', '地块名称', '地块面积/亩', '作物编号'])
    # print(land_types)
    # 将合并后的数据与 sheet11 合并，得到每个作物在每种土地类型上的亩产量
    sheet11 = sheet1.drop(columns=['序号', '作物编号'])
    yield_data = pd.merge(land_types, sheet11, left_on=['地块类型', '作物名称','种植季次'], right_on=['地块类型', '作物名称','种植季次'])
    # 计算每种作物的产量（种植面积 * 亩产量）
    yield_data['该地块产量'] = yield_data['种植面积/亩'] * yield_data['亩产量/斤']

    # 汇总每种作物在每一季度的总产量作为预测销售量
    expected_sales_volume = yield_data.groupby(['种植季次', '作物名称'])['该地块产量'].sum().reset_index()
    expected_sales_volume = expected_sales_volume.rename(columns={'该地块产量': '当季预期销售量'})
    expected_sales_volume.to_excel('预期销售量（季节）.xlsx', index=False)
    # print(expected_sales_volume)
    return sheet1, sheet2, sheet3, expected_sales_volume

# 销售单价处理函数，取平均值
def convert_price_range_to_average(price_range):
    try:
        # 分离区间的两个数值
        min_price, max_price = price_range.split('-')
        # 将字符串转换为浮点数
        min_price = float(min_price)
        max_price = float(max_price)
        # 计算平均数
        return (min_price + max_price) / 2
    except Exception as e:
        # 如果出现任何错误，返回 NaN
        print(f"Error processing {price_range}: {e}")
        return float('nan')

# 数据准备
def prepare_data(sheet_yield_and_price_2023, sheet_crop_planting_2023):
    # 数据处理
    crops = sheet_crop_planting_2023[['作物名称', '作物类型']]
    sheet_yield_and_price_2023['亩产量/斤'] = pd.to_numeric(sheet_yield_and_price_2023['亩产量/斤'], errors='coerce')
    sheet_yield_and_price_2023['种植成本/(元/亩)'] = pd.to_numeric(sheet_yield_and_price_2023['种植成本/(元/亩)'], errors='coerce')
    sheet_yield_and_price_2023['销售单价/(元/斤)'] = sheet_yield_and_price_2023['销售单价/(元/斤)'].apply(convert_price_range_to_average)
    fields = sheet_yield_and_price_2023[['地块类型', '作物名称', '种植季次', '亩产量/斤', '种植成本/(元/亩)', '销售单价/(元/斤)']]
    fields.to_excel('平均销售单价.xlsx', index=False)

    return crops, fields

# 创建单年模型
def define_model(solved_decision_vars,sheet_crop_planting_2023,sheet_fields_name_and_area, fields, expected_sales_volume, year):
    # 创建地块编号到地块类型的映射
    field_type_mapping = sheet_fields_name_and_area.set_index('地块名称')['地块类型'].to_dict()
    # 决策变量
    decision_vars = {}
    for _, row in fields.iterrows():
        field_type = row['地块类型']
        crop_name = row['作物名称']
        season = row['种植季次']
        # 迭代所有地块编号，并检查是否匹配
        for field_num, mapped_field_type in field_type_mapping.items():
            if mapped_field_type == field_type:
                    decision_vars[(field_num, field_type, crop_name, season, year)] = pulp.LpVariable(
                        f"area_{field_num}_{field_type}_{crop_name}_{season}_{year}",
                        lowBound=0, cat='Continuous'
                    )

    model = pulp.LpProblem(f"Maximize_Revenue_{year}", pulp.LpMaximize)

    # 目标函数：最大化该年的收益
    # 某种种植季次的产量和

    total_production = {}  # 分季总产量

    for crop_name in fields['作物名称'].unique():
        for season in fields[fields['作物名称'] == crop_name]['种植季次']:
                total_production[(crop_name, season, year)] = pulp.lpSum(
                    decision_vars.get((field_num, field_type, crop_name, season, year), 0)
                        * fields[(fields['地块类型'] == field_type) &
                                (fields['作物名称'] == crop_name) &
                                (fields['种植季次'] == season)]['亩产量/斤'].values[0]
                    for field_type in fields[(fields['作物名称']==crop_name) & (fields['种植季次'] == season)]['地块类型']
                    for field_num,field_type_value in field_type_mapping.items()
                    if field_type_mapping.get(field_num) == field_type
                )

    # 总种植成本
    total_cost = pulp.lpSum([
        decision_vars[(field_num, row['地块类型'], row['作物名称'], row['种植季次'], year)] * row['种植成本/(元/亩)']
        for _, row in fields.iterrows()
        for field_num, mapped_field_type in field_type_mapping.items()
        if mapped_field_type == row['地块类型']
    ])

    # 新的一堆决策变量
    under_exp_sv = {}
    beyond_exp_sv = {}
    for crop_name in fields['作物名称'].unique():
        for season in fields[fields['作物名称'] == crop_name]['种植季次'].unique():
            # 创建新的决策变量
            under_exp_sv[(crop_name,season)] = pulp.LpVariable(f'under_exp_sv_{crop_name}_{season}', lowBound=0, cat='Continuous')
            beyond_exp_sv[(crop_name,season)] = pulp.LpVariable(f'beyond_exp_sv_{crop_name}_{season}', lowBound=0, cat='Continuous')
            exp_sv = expected_sales_volume[(expected_sales_volume['种植季次'] == season)&(expected_sales_volume['作物名称'] == crop_name)]['当季预期销售量'].values[0]

            # 添加约束：对每个季节的每种作物都要添加
            # TODO:未达预期销售量时，beyond为负数。对1.1无影响
            model += under_exp_sv[(crop_name,season)] + beyond_exp_sv[(crop_name,season)] == total_production[(crop_name, season, year)]
            model += under_exp_sv[(crop_name,season)] <= exp_sv


    # 总销售收益
    # 超过预期销售量的部分无法售出
    # TODO: 1.2修改处
    total_revenue = pulp.lpSum([
        under_exp_sv[(crop_name,season)]
        *fields[
        (fields['种植季次'] == season) &
        (fields['作物名称'] == crop_name)
        ]['销售单价/(元/斤)'].values[0]

        for crop_name in fields['作物名称'].unique()
        for season in fields[fields['作物名称'] == crop_name]['种植季次']
    ]) - total_cost

    model += total_revenue, f"Total_Revenue_{year}"

    # 约束条件
    # 1. 每种作物在同一地块（含大棚）都不能连续重茬种植
    # 1.1 同一年的第一季、第二季。仅限于智慧大棚中全部作物。
    smart_greenhouse_field_nums = sheet_fields_name_and_area[sheet_fields_name_and_area['地块类型'] == '智慧大棚']['地块名称']
    for field_num in smart_greenhouse_field_nums:
        for crop_name in fields[fields['地块类型'] == '智慧大棚']['作物名称'].unique():
            area_first_season = decision_vars.get((field_num, '智慧大棚', crop_name, '第一季', year), 0)
            area_second_season = decision_vars.get((field_num, '智慧大棚', crop_name, '第二季', year), 0)
            # 添加约束：保证至少有一个季节的种植面积为0
            model += (area_first_season == 0 or area_second_season == 0)

    # 1.2 单季作物的连续两年。仅限单季作物
    for field_type in ['平旱地', '梯田', '山坡地', '水浇地']:
        field_nums = sheet_fields_name_and_area[sheet_fields_name_and_area['地块类型'] == field_type]['地块名称']
        for field_num in field_nums:
            for crop_name in fields[(fields['地块类型'] == field_type) & (fields['种植季次'] == '单季')]['作物名称'].unique():
                if year == 2024:
                    area_this_year = decision_vars.get((field_num, field_type, crop_name, '单季', year), 0)  # 水浇地中的非水稻作物这里返回0
                    # 查找2023年对应地块对应作物的种植量
                    filtered_data = sheet_crop_planting_2023[
                        (sheet_crop_planting_2023['作物名称'] == crop_name) &
                        (sheet_crop_planting_2023['种植地块'] == field_num)
                        ]['种植面积/亩']

                    if not filtered_data.empty:
                        model += (area_this_year == 0)
                else:
                    area_last_year = solved_decision_vars.get(f"area_{field_num}_{field_type}_{crop_name}_单季_{year - 1}")
                    if area_last_year is not None and area_last_year > 0:
                        area_this_year = decision_vars.get((field_num, field_type, crop_name, '单季', year), 0)
                        model += (area_this_year == 0)

    # 1.3 上一年的第二季与下一年的第一季不能连续种植。仅需考虑智慧大棚中的作物

    for field_num in smart_greenhouse_field_nums:
        for crop_name in fields[fields['地块类型'] == '智慧大棚']['作物名称'].unique():
            if year == 2024:
                area_this_first_season = decision_vars.get((field_num, '智慧大棚', crop_name, '第一季', year), 0)
                # 查找2023年对应地块对应作物的种植量
                filtered_data = sheet_crop_planting_2023[
                    (sheet_crop_planting_2023['作物名称'] == crop_name) &
                    (sheet_crop_planting_2023['种植地块'] == field_num)
                    ]['种植面积/亩']

                if not filtered_data.empty:
                    model += (area_this_first_season == 0)

            else:
                area_last_second_season = solved_decision_vars.get(f"area_{field_num}_智慧大棚_{crop_name}_第二季_{year - 1}")
                if area_last_second_season is not None and area_last_second_season > 0:
                    area_this_first_season = decision_vars.get((field_num, '智慧大棚', crop_name, '第一季', year), 0)
                    model += (area_this_first_season == 0)


    # 2. 每个地块（含大棚）的所有土地三年内至少种植一次豆类作物
    if year != 2024:
        bean_crops = ['黄豆', '黑豆','红豆','绿豆','爬豆','豇豆','刀豆','芸豆']  # 豆类作物名称
        for field_type in['平旱地', '梯田', '山坡地', '水浇地','普通大棚','智慧大棚']:
                field_nums = sheet_fields_name_and_area[sheet_fields_name_and_area['地块类型'] == field_type]['地块名称']
                for field_num in field_nums:
                    # 前两年该地块上的豆类产量
                    bean_volume_in_last_2year = 0
                    if year == 2025:
                        for crop_name in bean_crops:
                            for season in fields[fields['作物名称'] == crop_name]['种植季次'].unique():
                                if solved_decision_vars.get(
                                        f"area_{field_num}_{field_type}_{crop_name}_{season}_{year - 1}") is not None:
                                    bean_volume_in_last_2year += solved_decision_vars.get(f"area_{field_num}_{field_type}_{crop_name}_{season}_{year - 1}")

                                filtered_data = sheet_crop_planting_2023[
                                    (sheet_crop_planting_2023['作物名称'] == crop_name) &
                                    (sheet_crop_planting_2023['种植地块'] == field_num)
                                    ]['种植面积/亩']
                                if not filtered_data.empty:
                                    bean_volume_in_last_2year += filtered_data.values[0]

                    elif year > 2025:
                        for crop_name in bean_crops:
                            for season in fields[fields['作物名称'] == crop_name]['种植季次'].unique():
                                if solved_decision_vars.get(
                                        f"area_{field_num}_{field_type}_{crop_name}_{season}_{year - 1}") is not None:
                                    bean_volume_in_last_2year += solved_decision_vars.get(f"area_{field_num}_{field_type}_{crop_name}_{season}_{year - 1}")
                                if solved_decision_vars.get(
                                        f"area_{field_num}_{field_type}_{crop_name}_{season}_{year - 2}") is not None:
                                    bean_volume_in_last_2year += solved_decision_vars.get(f"area_{field_num}_{field_type}_{crop_name}_{season}_{year - 2}")
                    if bean_volume_in_last_2year == 0:
                        model += pulp.lpSum([
                            decision_vars.get((field_num,field_type, crop_name, season ,year), 0)
                            for crop_name in bean_crops
                            for season in fields[fields['作物名称'] == crop_name]['种植季次'].unique()
                        ]) >= 0.06

    # 4. 某地块某季节的种植总面积小于该地块面积
    for season in ['第一季', '第二季','单季']:
        for field_type in fields['地块类型'].unique():
            field_nums = sheet_fields_name_and_area[sheet_fields_name_and_area['地块类型'] == field_type]['地块名称']
            for field_num in field_nums:
                    area_of_this_fields = sheet_fields_name_and_area[sheet_fields_name_and_area['地块名称']==field_num]['地块面积/亩']
                    model += (pulp.lpSum([
                                decision_vars.get((field_num, field_type, crop_name, season, year), 0)
                                for crop_name in fields[fields['地块类型'] == field_type]['作物名称'].unique()
                            ]) <= area_of_this_fields)

    # 5. 水浇地单季（只有水稻）与双季不能共存
    field_nums = sheet_fields_name_and_area[sheet_fields_name_and_area['地块类型'] == '水浇地']['地块名称']
    for field_num in field_nums:
        sjd_single_area = decision_vars.get((field_num, '水浇地', '水稻', '单季', year), 0)
        for crop_name in fields[(fields['地块类型'] == '水浇地')&(fields['种植季次'] != '单季')]['作物名称'].unique():
            sjd_first_area = decision_vars.get((field_num,'水浇地', crop_name, '第一季', year),0)
            sjd_second_area = decision_vars.get((field_num,'水浇地', crop_name, '第二季', year),0)
            model += (sjd_single_area == 0 or (sjd_first_area == 0 and sjd_second_area == 0))

    return model, decision_vars


# 求解模型
def solve_model(model):
    model.solve()
    return model


# 保存结果到Excel
def save_results(decision_vars,results,year):
    for var in decision_vars.values():
        results.append({
            '地块名称': var.name.split('_')[1],
            '地块类型': var.name.split('_')[2],
            '作物名称': var.name.split('_')[3],
            '季次': var.name.split('_')[4],
            '年份': var.name.split('_')[5],
            '种植面积': var.varValue
        })
    if year == 2030:
        result_df = pd.DataFrame(results)
        result_df.to_excel("my_result1_1.xlsx", index=False)


def process_file(my_result_file, cache_files, attachment_file):
    # 解包 cache_files
    cache_file_blank, cache_file_final = cache_files

    # Part 1: 读取 my_result.xlsx 文件并更新缓存文件
    # 读取 my_result.xlsx 文件
    my_result_df = pd.read_excel(my_result_file)

    # 替换季次中的"单季"为"第一季"
    my_result_df['季次'] = my_result_df['季次'].replace('单季', '第一季')

    # 打开缓存文件中的所有 sheet
    xls = pd.ExcelFile(cache_file_blank)
    years_sheets = {sheet_name: pd.read_excel(xls, sheet_name=sheet_name) for sheet_name in xls.sheet_names}

    # 作物列的对应关系
    crop_columns = {
        '黄豆': '黄豆', '黑豆': '黑豆', '红豆': '红豆', '绿豆': '绿豆', '爬豆': '爬豆',
        '小麦': '小麦', '玉米': '玉米', '谷子': '谷子', '高粱': '高粱', '黍子': '黍子',
        '荞麦': '荞麦', '南瓜': '南瓜', '红薯': '红薯', '莜麦': '莜麦', '大麦': '大麦',
        '水稻': '水稻', '豇豆': '豇豆', '刀豆': '刀豆', '芸豆': '芸豆', '土豆': '土豆',
        '西红柿': '西红柿', '茄子': '茄子', '菠菜': '菠菜', '青椒': '青椒', '菜花': '菜花',
        '包菜': '包菜', '油麦菜': '油麦菜', '小青菜': '小青菜', '黄瓜': '黄瓜', '生菜': '生菜',
        '辣椒': '辣椒', '空心菜': '空心菜', '黄心菜': '黄心菜', '芹菜': '芹菜', '大白菜': '大白菜',
        '白萝卜': '白萝卜', '红萝卜': '红萝卜', '榆黄菇': '榆黄菇', '香菇': '香菇', '白灵菇': '白灵菇',
        '羊肚菌': '羊肚菌'
    }

    # 遍历 my_result.xlsx 中的数据并将其写入到缓存文件
    for index, row in my_result_df.iterrows():
        field = row['地块名称']
        crop = row['作物名称']
        season = row['季次']
        year = row['年份']
        area = row['种植面积']

        # 1. 找年份对应的 sheet
        sheet_name = str(year)
        if sheet_name in years_sheets:
            sheet_df = years_sheets[sheet_name]

            # 2. 找到对应的季次 ("第一季" 或 "第二季")
            season_rows = sheet_df[sheet_df['季次'] == season]
            if season_rows.empty:
                print(f"Season '{season}' not found in sheet {sheet_name}")
                continue

            # 3. 找到对应的地块
            field_row = season_rows[season_rows['地块名'] == field]
            if field_row.empty:
                print(f"Field '{field}' not found in season '{season}' in sheet {sheet_name}")
                continue

            # 获取地块所在的行索引
            field_row_idx = field_row.index[0]

            # 4. 找到作物列
            if crop in crop_columns:
                crop_col = crop_columns[crop]
            else:
                print(f"Crop '{crop}' not found in crop_columns")
                continue

            # 确保作物列存在
            if crop_col not in sheet_df.columns:
                print(f"Crop column '{crop_col}' not found in sheet '{sheet_name}'")
                continue

            # 5. 写入种植面积数据
            if pd.isna(sheet_df.at[field_row_idx, crop_col]):  # 如果当前单元格为空
                sheet_df.at[field_row_idx, crop_col] = area

    # 保存更新后的缓存文件为新的文件
    with pd.ExcelWriter(cache_file_final, engine='openpyxl', mode='w') as writer:
        for sheet_name, df in years_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Data update complete in {cache_file_final}.")

    # Part 2: 将缓存文件中的数据粘贴到 附件文件中

    # 打开缓存文件
    a_workbook = openpyxl.load_workbook(cache_file_final)

    # 打开附件文件
    b_workbook = openpyxl.load_workbook(attachment_file)

    # 遍历缓存文件和附件文件的所有 sheet
    for sheet_name in a_workbook.sheetnames:
        # 取出缓存文件和附件文件中对应的 sheet
        a_sheet = a_workbook[sheet_name]
        b_sheet = b_workbook[sheet_name]

        # 提取缓存文件 C2 到 AQ83 的数据
        for row_idx in range(2, 84):  # 行范围从 2 到 83
            for col_idx in range(3, 44):  # 列范围从 C (3) 到 AQ (43)
                # 获取缓存文件中的数据
                a_value = a_sheet.cell(row=row_idx, column=col_idx).value

                # 将数据粘贴到附件文件相应单元格中，保持原有格式
                b_sheet.cell(row=row_idx, column=col_idx).value = a_value

    # 保存修改后的附件文件，保留原有格式和合并单元格
    b_workbook.save(attachment_file)

    print(
        f"Data successfully copied from {cache_file_final} to {attachment_file}, preserving all formatting and merged cells.")



# 主流程
def main():
    sheet_yield_and_price_2023, sheet_crop_planting_2023, sheet_fields_name_and_area, expected_sales_volume = read_data()
    crops, fields = prepare_data(sheet_yield_and_price_2023, sheet_crop_planting_2023)
    print(crops, fields)
    expected_sales_volume_1 = pd.read_excel('datas/预期每季销售量（补充版）.xlsx')
    expected_sales_volume_1['作物名称'] = expected_sales_volume_1['作物名称'].str.strip() # 去空格

    models = {}
    solved_decision_vars = {}   #已求解的decision_vars
    results = []
    for year in range(2024,2031):
        model, decision_vars = define_model(solved_decision_vars,sheet_crop_planting_2023, sheet_fields_name_and_area, fields, expected_sales_volume_1,year)
        models[year] = solve_model(model)   #求解
        save_results(decision_vars,results,year)

        # 遍历决策变量并存储它们的值
        for key, var in decision_vars.items():
            field_num, field_type, crop_name, season, _ = key
            var_name = f"area_{field_num}_{field_type}_{crop_name}_{season}_{year}"
            solved_decision_vars[var_name] = pulp.value(var)
            # 示例调用，每组文件逐个处理
        # 输出年利润
        objective_value = pulp.value(models[year].objective)
        print(f"{year}年利润: {objective_value}")

    process_file('my_result1_1.xlsx', ('datas/缓存-result1_1（空白）.xlsx', 'datas/缓存-result1_1.xlsx'),
                         'datas/附件3-result1_1.xlsx')



if __name__ == "__main__":
    main()
